"""Exportación Excel y reaplicación de reglas."""

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from tasa_estadistica.domain.tasa_estadistica_mapper import TasaEstadisticaMapper
from tasa_estadistica.export.excel_report import build_auditable_excel, reapply_mapper_sqlite
from tasa_estadistica.model.schemas import ConceptoLiquidacion, Liquidacion, RunParams
from tasa_estadistica.storage.sqlite_repo import SqliteRepo


def test_build_excel_and_reapply(tmp_path: Path) -> None:
    db = tmp_path / "t.db"
    repo = SqliteRepo(db)
    run = repo.start_run(
        RunParams(
            fecha_desde=date(2026, 1, 1),
            fecha_hasta=date(2026, 1, 31),
            cuit="20123456789",
            modo="mock",
        )
    )
    liq = Liquidacion(
        cuit="20123456789",
        id_externo="L1",
        numero="L1",
        conceptos=[],
    )
    repo.save_liquidaciones(run.run_id, [liq], TasaEstadisticaMapper())
    repo.close()

    xlsx = tmp_path / "out.xlsx"
    build_auditable_excel(db, xlsx)
    assert xlsx.is_file()

    n = reapply_mapper_sqlite(db)
    assert n >= 0


def test_build_excel_fecha_filter(tmp_path: Path) -> None:
    db = tmp_path / "t.db"
    repo = SqliteRepo(db)
    run = repo.start_run(
        RunParams(
            fecha_desde=date(2025, 5, 1),
            fecha_hasta=date(2025, 5, 31),
            cuit="20123456789",
            modo="mock",
        )
    )
    conc = ConceptoLiquidacion(
        codigo="1", descripcion="TASA DE ESTADISTICA", importe=Decimal("100")
    )
    liq_mayo = Liquidacion(
        cuit="20123456789",
        id_externo="L1",
        numero="L1",
        fecha=date(2025, 5, 15),
        conceptos=[conc],
    )
    liq_otro = Liquidacion(
        cuit="20123456789",
        id_externo="L2",
        numero="L2",
        fecha=date(2024, 1, 10),
        conceptos=[conc],
    )
    mapper = TasaEstadisticaMapper()
    repo.save_liquidaciones(run.run_id, [liq_mayo, liq_otro], mapper)
    repo.close()

    x_full = tmp_path / "full.xlsx"
    build_auditable_excel(db, x_full)
    wb_full = load_workbook(x_full)
    assert wb_full["detalle_conceptos"].max_row == 3
    assert wb_full["liquidaciones_cabecera"].max_row == 3

    x_mayo = tmp_path / "mayo.xlsx"
    build_auditable_excel(
        db, x_mayo, fecha_desde=date(2025, 5, 1), fecha_hasta=date(2025, 5, 31)
    )
    wb_mayo = load_workbook(x_mayo)
    assert wb_mayo["detalle_conceptos"].max_row == 2
    assert wb_mayo["liquidaciones_cabecera"].max_row == 2
    meta = wb_mayo["meta_export"]
    assert meta["A2"].value == "fecha_filtro_desde"
    assert meta["B2"].value == "2025-05-01"
    assert meta["B3"].value == "2025-05-31"
