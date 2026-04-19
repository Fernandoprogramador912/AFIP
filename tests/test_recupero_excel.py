"""Export Excel modelo Recupero V2."""

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from tasa_estadistica.domain.tasa_estadistica_mapper import TasaEstadisticaMapper
from tasa_estadistica.export.recupero_excel import (
    RECUPERO_V2_HEADERS,
    build_recupero_v2_excel,
    recupero_resumen_local,
)
from tasa_estadistica.model.schemas import ConceptoLiquidacion, Liquidacion, RunParams
from tasa_estadistica.storage.sqlite_repo import SqliteRepo


def test_recupero_v2_columns_and_te_split(tmp_path: Path) -> None:
    db = tmp_path / "t.db"
    repo = SqliteRepo(db)
    run = repo.start_run(
        RunParams(
            fecha_desde=date(2025, 5, 1),
            fecha_hasta=date(2025, 5, 31),
            cuit="20123456789",
            modo="mock",
        )
    )
    raw = {
        "identificador_destinacion": "25001IC00000001X",
        "liquidacion_resumen": {
            "Cotizacion": "1100",
        },
        "declaracion_listado": {
            "NombreProveedorExterior": "PROVEEDOR TEST SA",
            "ValorFOB": "100",
            "ValorFlete": "200",
            "ValorSeguro": "50",
        },
    }
    liq = Liquidacion(
        cuit="20123456789",
        id_externo="x",
        numero="x",
        fecha=date(2025, 5, 10),
        destinacion_id="25001IC00000001X",
        conceptos=[
            ConceptoLiquidacion(
                codigo="011",
                descripcion="TASA",
                importe=Decimal("100"),
                moneda="ARS",
                raw={"MontoPagado": "100"},
            ),
            ConceptoLiquidacion(
                codigo="061",
                descripcion="MONT MAX",
                importe=Decimal("50"),
                moneda="ARS",
                raw={"MontoPagado": "50"},
            ),
        ],
        raw=raw,
    )
    repo.save_liquidaciones(run.run_id, [liq], TasaEstadisticaMapper())
    repo.close()

    out = tmp_path / "r.xlsx"
    build_recupero_v2_excel(db, out, date(2025, 5, 1), date(2025, 5, 31))
    assert (tmp_path / "r.csv").is_file()
    wb = load_workbook(out, read_only=True, data_only=True)
    ws = wb["recupero_V2"]
    assert ws["C3"].value == RECUPERO_V2_HEADERS[2]  # OFICIALIZACION header row 3
    # Fila 4 = primera fila de datos
    assert ws["A4"].value == "PROVEEDOR TEST SA"
    assert ws["B4"].value == "25001IC00000001X"
    assert ws["D4"].value == 1100.0
    assert ws["F4"].value == 100
    assert ws["G4"].value == 50
    assert ws["H4"].value == 0
    assert ws["I4"].value == 150
    # FOB=K, FLETE=L, SEGURO=M, CIF DOCUMENTAL=N (FOB+FLETE+SEGURO)
    assert ws["K4"].value == 100
    assert ws["L4"].value == 200
    assert ws["M4"].value == 50
    assert ws["N4"].value == 350
    wb.close()

    res = recupero_resumen_local(db, date(2025, 5, 1), date(2025, 5, 31))
    assert res.get("recupero_avisos_fob") == []
    assert res["n_despachos"] == 1
    assert res["total_tasa_recupero_ars"] == Decimal("150")
    assert res["n_liquidaciones_en_rango"] == 1
    assert res["n_destinaciones_distintas_en_rango"] == 1
    assert "IC" in res["destinacion_subcadenas"]


def test_recupero_prefiere_fila_moa_mismo_destino_que_complemento_zeep(tmp_path: Path) -> None:
    """
    MOA + complemento Zeep insertan dos liquidaciones (id_externo distinto) con el mismo
    destinacion_id; debe usarse la fila MOA para FOB aunque el Zeep tenga id mayor.
    """
    db = tmp_path / "dup.db"
    repo = SqliteRepo(db)
    run = repo.start_run(
        RunParams(
            fecha_desde=date(2025, 6, 1),
            fecha_hasta=date(2025, 6, 30),
            cuit="20123456789",
            modo="mock",
        )
    )
    dest = "25001IC00009999Z"
    raw_moa = {
        "identificador_destinacion": dest,
        "moa_detallada_caratula": {
            "Caratula": {
                "MontoFobTotal": "999",
                "MontoFleteTotal": "10",
                "MontoSeguroTotal": "1",
            }
        },
    }
    raw_zeep = {"identificador_destinacion": dest, "fuente": "complemento_zeep"}
    conceptos = [
        ConceptoLiquidacion(
            codigo="011",
            descripcion="TASA",
            importe=Decimal("50"),
            moneda="ARS",
            raw={},
        ),
    ]
    liq_moa = Liquidacion(
        cuit="20123456789",
        id_externo=f"{dest}:LIQ1",
        numero="LIQ1",
        fecha=date(2025, 6, 15),
        destinacion_id=dest,
        conceptos=conceptos,
        raw=raw_moa,
    )
    liq_zeep = Liquidacion(
        cuit="20123456789",
        id_externo="AFIP-999",
        numero="999",
        fecha=date(2025, 6, 15),
        destinacion_id=dest,
        conceptos=conceptos,
        raw=raw_zeep,
    )
    repo.save_liquidaciones(run.run_id, [liq_moa, liq_zeep], TasaEstadisticaMapper())
    repo.close()

    res = recupero_resumen_local(db, date(2025, 6, 1), date(2025, 6, 30))
    assert res["n_despachos"] == 1
    row = res["filas"][0]
    assert row[RECUPERO_V2_HEADERS.index("FOB")] == 999
    assert row[RECUPERO_V2_HEADERS.index("FLETE")] == 10
    assert row[RECUPERO_V2_HEADERS.index("SEGURO")] == 1


def test_recupero_merge_fob_otra_guarda_mismo_di(tmp_path: Path) -> None:
    """
    Refetch: la fila más nueva puede tener bloques MOA vacíos pero otra guarda del mismo
    D.I. conserva MontoFobTotal; se unifica al armar la fila recupero.
    """
    db = tmp_path / "merge.db"
    repo = SqliteRepo(db)
    run = repo.start_run(
        RunParams(
            fecha_desde=date(2025, 7, 1),
            fecha_hasta=date(2025, 7, 31),
            cuit="20123456789",
            modo="mock",
        )
    )
    dest = "25001IC00007777Y"
    raw_viejo = {
        "moa_detallada_caratula": {
            "Caratula": {
                "MontoFobTotal": "500",
                "MontoFleteTotal": "20",
                "MontoSeguroTotal": "5",
            }
        },
    }
    raw_nuevo = {
        "moa_detallada_caratula": {"Caratula": {}},
        "moa_detallada_liquidaciones_detalle": {},
    }
    conceptos = [
        ConceptoLiquidacion(
            codigo="011",
            descripcion="TASA",
            importe=Decimal("10"),
            moneda="ARS",
            raw={},
        ),
    ]
    liq_old = Liquidacion(
        cuit="20123456789",
        id_externo=f"{dest}:OLD",
        numero="OLD",
        fecha=date(2025, 7, 10),
        destinacion_id=dest,
        conceptos=conceptos,
        raw=raw_viejo,
    )
    liq_new = Liquidacion(
        cuit="20123456789",
        id_externo=f"{dest}:NEW",
        numero="NEW",
        fecha=date(2025, 7, 12),
        destinacion_id=dest,
        conceptos=conceptos,
        raw=raw_nuevo,
    )
    repo.save_liquidaciones(run.run_id, [liq_old, liq_new], TasaEstadisticaMapper())
    repo.close()

    res = recupero_resumen_local(db, date(2025, 7, 1), date(2025, 7, 31))
    assert res["n_despachos"] == 1
    row = res["filas"][0]
    assert row[RECUPERO_V2_HEADERS.index("FOB")] == 500
    assert row[RECUPERO_V2_HEADERS.index("FLETE")] == 20
    assert row[RECUPERO_V2_HEADERS.index("SEGURO")] == 5
