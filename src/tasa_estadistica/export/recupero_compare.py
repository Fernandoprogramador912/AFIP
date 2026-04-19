"""
Comparación fila a fila: hoja modelo Excel (V2_Ejemplo) vs filas generadas desde SQLite.

Útil para ver por qué el panel no coincide con el libro: mismas fórmulas en Python,
pero insumos distintos (AFIP) o celdas del Excel con valores cacheados distintos.
"""

from __future__ import annotations

import sqlite3
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tasa_estadistica.config.settings import Settings, get_settings
from tasa_estadistica.export.recupero_excel import RECUPERO_V2_HEADERS, recupero_v2_data_rows


def _norm_scalar(v: Any) -> Any:
    """Normaliza para comparar Excel vs Python."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        t = v.strip()
        return "" if t == "" else t
    if isinstance(v, (int, float)):
        d = Decimal(str(v))
        if d == d.to_integral():
            return int(d)
        return float(d)
    return v


def _same_cell(excel_v: Any, py_v: Any) -> bool:
    a, b = _norm_scalar(excel_v), _norm_scalar(py_v)
    if a == b:
        return True
    if (a is None or a == "") and (b is None or b == ""):
        return True
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        da, db = Decimal(str(a)), Decimal(str(b))
        diff = abs(da - db)
        tol = max(Decimal("1e-6"), abs(db) * Decimal("1e-9"))
        return diff <= tol
    return False


def read_excel_recupero_v2_data_rows(
    excel_path: Path,
    *,
    sheet_name: str = "V2_Ejemplo",
    data_start_row: int = 4,
    ncols: int = 25,
) -> list[tuple[str, list[Any]]]:
    """
    Lee filas de datos del modelo (columnas A.. hasta ncols).

    Devuelve lista de (destinacion_id, celdas[0..ncols-1]).
    Corta cuando la columna B (D.I.) está vacía.
    """
    path = Path(excel_path)
    if not path.is_file():
        raise FileNotFoundError(str(path))

    out: list[tuple[str, list[Any]]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"No existe la hoja {sheet_name!r}. Hojas: {wb.sheetnames!r}"
            )
        ws = wb[sheet_name]
        # read_only: max_row puede ser poco fiable; cortamos por columna B vacía.
        for r in range(data_start_row, data_start_row + 10000):
            b = ws.cell(row=r, column=2).value
            dest = str(b).strip() if b is not None else ""
            if not dest:
                break
            row = [ws.cell(row=r, column=c).value for c in range(1, ncols + 1)]
            out.append((dest, row))
    finally:
        wb.close()
    return out


def compare_excel_vs_sqlite(
    excel_path: Path,
    sqlite_path: Path,
    fecha_desde: date,
    fecha_hasta: date,
    *,
    sheet_name: str = "V2_Ejemplo",
    data_start_row: int = 4,
    settings: Settings | None = None,
) -> list[dict[str, Any]]:
    """
    Para cada fila del Excel con D.I., busca la misma destinación en SQLite y compara 25 columnas.

    Retorna lista de dicts: destinacion_id, status ('ok'|'solo_excel'|'diff'), diffs, message.
    """
    excel_rows = read_excel_recupero_v2_data_rows(
        excel_path, sheet_name=sheet_name, data_start_row=data_start_row
    )

    settings = settings or get_settings()
    conn = sqlite3.connect(sqlite_path)
    conn.row_factory = sqlite3.Row
    try:
        py_rows = recupero_v2_data_rows(conn, fecha_desde, fecha_hasta, settings)
    finally:
        conn.close()

    by_dest: dict[str, list[Any]] = {}
    for row in py_rows:
        if len(row) < 2:
            continue
        d = str(row[1] or "").strip()
        if d:
            by_dest[d] = row

    results: list[dict[str, Any]] = []
    excel_dests = {d for d, _ in excel_rows}

    for dest, xrow in excel_rows:
        pyrow = by_dest.get(dest)
        if pyrow is None:
            results.append(
                {
                    "destinacion_id": dest,
                    "status": "solo_excel",
                    "message": (
                        "Este D.I. no aparece en SQLite para el rango de fechas "
                        "(o no hay liquidación IC reciente). "
                        "El panel solo muestra lo cargado con fetch."
                    ),
                    "diffs": [],
                }
            )
            continue

        diffs: list[dict[str, Any]] = []
        for i, name in enumerate(RECUPERO_V2_HEADERS):
            xv = xrow[i] if i < len(xrow) else None
            pv = pyrow[i] if i < len(pyrow) else None
            if not _same_cell(xv, pv):
                diffs.append({"columna": name, "excel": xv, "python": pv})

        results.append(
            {
                "destinacion_id": dest,
                "status": "ok" if not diffs else "diff",
                "diffs": diffs,
                "message": "",
            }
        )

    sqlite_only = set(by_dest.keys()) - excel_dests
    if sqlite_only:
        muestra = ", ".join(sorted(sqlite_only)[:8])
        extra = "…" if len(sqlite_only) > 8 else ""
        results.append(
            {
                "destinacion_id": "*",
                "status": "info",
                "message": f"En SQLite hay {len(sqlite_only)} D.I. que no están en el Excel "
                f"(filas de datos leídas). Ej.: {muestra}{extra}",
                "diffs": [],
            }
        )

    return results


def format_compare_report(results: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    for r in results:
        did = r["destinacion_id"]
        st = r["status"]
        if st == "ok":
            lines.append(f"[OK] {did} — mismas 25 columnas (tras normalizar números/fechas).")
            continue
        if st == "solo_excel":
            lines.append(f"[SOLO EXCEL] {did}")
            lines.append(f"  {r['message']}")
            continue
        if st == "info":
            lines.append(f"[INFO] {r['message']}")
            continue
        lines.append(f"[DIFERENCIAS] {did}")
        for d in r["diffs"]:
            lines.append(
                f"  · {d['columna']}: Excel={d['excel']!r} | Python={d['python']!r}"
            )
    return "\n".join(lines) if lines else "(sin filas para comparar)"
