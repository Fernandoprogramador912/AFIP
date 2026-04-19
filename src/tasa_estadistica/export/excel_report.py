"""Exportación Excel auditable: detalle, resumen tasa, trazabilidad."""

from __future__ import annotations

import json
import sqlite3
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from tasa_estadistica.config.settings import get_settings, get_tasa_mapper_from_settings
from tasa_estadistica.domain.tasa_estadistica_mapper import TasaEstadisticaMapper
from tasa_estadistica.export.excel_raw_flat import (
    CO_FLAT_KEYS,
    LIQ_FLAT_KEYS,
    flatten_concepto_raw_json,
    flatten_liquidacion_raw_json,
    importe_concepto_efectivo,
)
from tasa_estadistica.report.ic_tasa_report import query_ic_tasa_rows, total_monto


def _dec(s: str | None) -> Decimal:
    if not s:
        return Decimal("0")
    return Decimal(str(s).replace(",", "."))


def _fecha_filtro_sql(
    fecha_desde: date | None, fecha_hasta: date | None
) -> tuple[str, list[str]]:
    """Filtro por `liquidaciones.fecha` (ISO en SQLite). Ambas fechas o ninguna."""
    if fecha_desde is None and fecha_hasta is None:
        return "", []
    if fecha_desde is None or fecha_hasta is None:
        raise ValueError("fecha_desde y fecha_hasta deben indicarse juntas")
    return "AND l.fecha >= ? AND l.fecha <= ?", [
        fecha_desde.isoformat(),
        fecha_hasta.isoformat(),
    ]


def build_auditable_excel(
    db_path: Path,
    output_path: Path,
    fecha_desde: date | None = None,
    fecha_hasta: date | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    fc_sql, fc_params = _fecha_filtro_sql(fecha_desde, fecha_hasta)

    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT r.run_id, r.started_at, r.params_json, r.modo,
               l.cuit, l.id_externo, l.numero, l.fecha, l.destinacion_id,
               l.raw_json AS liquidacion_raw_json,
               c.codigo, c.descripcion, c.importe, c.moneda,
               c.raw_json AS concepto_raw_json,
               c.es_tasa_estadistica, c.match_score, c.match_reason
        FROM conceptos_liquidacion c
        JOIN liquidaciones l ON l.id = c.liquidacion_id
        JOIN extraction_runs r ON r.run_id = c.run_id
        WHERE 1=1
        {fc_sql}
        ORDER BY r.started_at DESC, l.id, c.id
        """,
        fc_params,
    )
    rows = [dict(x) for x in cur.fetchall()]
    for r in rows:
        fl = flatten_liquidacion_raw_json(r.get("liquidacion_raw_json"))
        fc = flatten_concepto_raw_json(r.get("concepto_raw_json"))
        r["importe_efectivo"] = str(
            importe_concepto_efectivo(r.get("importe"), r.get("concepto_raw_json"))
        )
        r.update(fl)
        r.update(fc)

    cur_liq = conn.cursor()
    cur_liq.execute(
        f"""
        SELECT r.run_id, r.started_at, r.modo, r.params_json,
               l.id AS liquidacion_id, l.cuit, l.id_externo, l.numero, l.fecha, l.destinacion_id,
               l.raw_json AS liquidacion_raw_json
        FROM liquidaciones l
        JOIN extraction_runs r ON r.run_id = l.run_id
        WHERE 1=1
        {fc_sql}
        ORDER BY r.started_at DESC, l.id
        """,
        fc_params,
    )
    rows_liq = [dict(x) for x in cur_liq.fetchall()]
    for r in rows_liq:
        r.update(flatten_liquidacion_raw_json(r.get("liquidacion_raw_json")))

    conn.close()

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_det = wb.active
    ws_det.title = "detalle_conceptos"
    detalle_headers = [
        "run_id",
        "started_at",
        "modo",
        "params_json",
        "cuit",
        "id_externo_liquidacion",
        "numero_liquidacion",
        "fecha_liquidacion",
        "destinacion_id",
        "codigo_concepto",
        "descripcion_concepto",
        "importe",
        "importe_efectivo",
        "moneda",
        "es_tasa_estadistica",
        "match_score",
        "match_reason",
    ]
    detalle_headers.extend(LIQ_FLAT_KEYS)
    detalle_headers.extend(CO_FLAT_KEYS)
    ws_det.append(detalle_headers)
    for r in rows:
        base = [
            r["run_id"],
            r["started_at"],
            r["modo"],
            r.get("params_json"),
            r["cuit"],
            r["id_externo"],
            r["numero"],
            r["fecha"],
            r["destinacion_id"],
            r["codigo"],
            r["descripcion"],
            r["importe"],
            r["importe_efectivo"],
            r["moneda"],
            r["es_tasa_estadistica"],
            r["match_score"],
            r["match_reason"],
        ]
        fl = flatten_liquidacion_raw_json(r.get("liquidacion_raw_json"))
        fc = flatten_concepto_raw_json(r.get("concepto_raw_json"))
        base.extend(fl[k] for k in LIQ_FLAT_KEYS)
        base.extend(fc[k] for k in CO_FLAT_KEYS)
        ws_det.append(base)

    # Resumen tasa
    ws_res = wb.create_sheet("resumen_tasa")
    ws_res.append(["cuit", "periodo", "destinacion_id", "suma_tasa_estadistica_ars", "filas"])
    by_key: dict[tuple[str, str, str | None], list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        if not r["es_tasa_estadistica"]:
            continue
        params = {}
        try:
            params = json.loads(r["params_json"] or "{}")
        except json.JSONDecodeError:
            pass
        periodo = f"{params.get('fecha_desde', '')}_{params.get('fecha_hasta', '')}"
        key = (r["cuit"], periodo, r["destinacion_id"])
        by_key[key].append(r)

    for (cuit, periodo, dest), items in sorted(by_key.items(), key=lambda x: (x[0][0], x[0][1])):
        total = sum(
            _dec(x.get("importe_efectivo"))
            for x in items
            if (x["moneda"] or "ARS").upper() == "ARS"
        )
        ws_res.append([cuit, periodo, dest or "", float(total), len(items)])

    # Trazabilidad
    ws_tr = wb.create_sheet("trazabilidad")
    ws_tr.append(["run_id", "started_at", "modo", "params_json", "registros_exportados"])
    runs: dict[str, dict[str, Any]] = {}
    for r in rows:
        rid = r["run_id"]
        if rid not in runs:
            runs[rid] = {
                "run_id": rid,
                "started_at": r["started_at"],
                "modo": r["modo"],
                "params_json": r["params_json"],
                "n": 0,
            }
        runs[rid]["n"] += 1
    for rid, meta in sorted(runs.items(), key=lambda x: x[1]["started_at"], reverse=True):
        ws_tr.append([rid, meta["started_at"], meta["modo"], meta["params_json"], meta["n"]])

    # Una fila por liquidación: cabecera + totales AFIP (mismo despacho, sin repetir por concepto)
    ws_liq = wb.create_sheet("liquidaciones_cabecera")
    liq_headers = [
        "run_id",
        "started_at",
        "modo",
        "params_json",
        "liquidacion_id",
        "cuit",
        "id_externo_liquidacion",
        "numero_liquidacion",
        "fecha_liquidacion",
        "destinacion_id",
    ]
    liq_headers.extend(LIQ_FLAT_KEYS)
    ws_liq.append(liq_headers)
    for r in rows_liq:
        row = [
            r["run_id"],
            r["started_at"],
            r["modo"],
            r.get("params_json"),
            r["liquidacion_id"],
            r["cuit"],
            r["id_externo"],
            r["numero"],
            r["fecha"],
            r["destinacion_id"],
        ]
        fl = flatten_liquidacion_raw_json(r.get("liquidacion_raw_json"))
        row.extend(fl[k] for k in LIQ_FLAT_KEYS)
        ws_liq.append(row)

    # Tasa estadística por despacho importación (mismo rango que el filtro, o todo histórico en BD)
    ws_ic = wb.create_sheet("tasa_ic_por_despacho")
    s = get_settings()
    m = get_tasa_mapper_from_settings(s)
    ic_desde = fecha_desde if fecha_desde is not None else date(2019, 1, 1)
    ic_hasta = fecha_hasta if fecha_hasta is not None else date(2035, 12, 31)
    ic_rows = query_ic_tasa_rows(db_path, ic_desde, ic_hasta, m.codigos, settings=s)
    ws_ic.append(
        [
            "destinacion_id",
            "fecha",
            "cuit",
            "numero",
            "codigo_tasa",
            "monto_tasa_estadistica",
            "match_reason",
        ]
    )
    for ir in ic_rows:
        monto = ir.get("monto_tasa_estadistica")
        if isinstance(monto, Decimal):
            monto = float(monto)
        ws_ic.append(
            [
                ir.get("destinacion_id"),
                ir.get("fecha"),
                ir.get("cuit"),
                ir.get("numero"),
                ir.get("codigo_tasa"),
                monto,
                ir.get("match_reason"),
            ]
        )
    tot = total_monto(ic_rows)
    ws_ic.append(
        [
            "TOTAL",
            "",
            "",
            "",
            "",
            float(tot) if isinstance(tot, Decimal) else tot,
            f"despachos={len(ic_rows)}",
        ]
    )

    ws_meta = wb.create_sheet("meta_export")
    ws_meta.append(["clave", "valor"])
    ws_meta.append(
        [
            "fecha_filtro_desde",
            fecha_desde.isoformat() if fecha_desde else "(sin filtro)",
        ]
    )
    ws_meta.append(
        [
            "fecha_filtro_hasta",
            fecha_hasta.isoformat() if fecha_hasta else "(sin filtro)",
        ]
    )
    ws_meta.append(
        [
            "nota",
            "Si indicó --desde/--hasta, solo liquidaciones en ese rango por fecha_liquidacion.",
        ]
    )

    # Ajuste simple de anchos
    for ws in (ws_det, ws_res, ws_tr, ws_liq, ws_ic, ws_meta):
        for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1), start=1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    wb.save(output_path)
    return output_path


def reapply_mapper_sqlite(db_path: Path, mapper: TasaEstadisticaMapper | None = None) -> int:
    """Recalcula flags de tasa en SQLite según reglas actuales del mapper."""
    m = mapper or TasaEstadisticaMapper()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT id, codigo, descripcion FROM conceptos_liquidacion")
    updated = 0
    for row in cur.fetchall():
        from tasa_estadistica.model.schemas import ConceptoLiquidacion

        conc = ConceptoLiquidacion(codigo=row["codigo"] or "", descripcion=row["descripcion"] or "")
        match = m.match_concepto(conc)
        cur.execute(
            """
            UPDATE conceptos_liquidacion
            SET es_tasa_estadistica = ?, match_score = ?, match_reason = ?
            WHERE id = ?
            """,
            (1 if match.matched else 0, match.score, match.reason, row["id"]),
        )
        updated += 1
    conn.commit()
    conn.close()
    return updated
