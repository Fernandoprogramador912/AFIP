"""Export tipo modelo «Recupero Tasa» (hoja V2): columnas alineadas al Excel de referencia."""

from __future__ import annotations

import csv
import json
import sqlite3
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from tasa_estadistica.config.settings import Settings, get_settings
from tasa_estadistica.export.excel_raw_flat import (
    extract_proveedor_from_liquidacion_raw_json,
    extract_recupero_valores_extra,
    flatten_liquidacion_raw_json,
    importe_concepto_efectivo,
)
from tasa_estadistica.export.recupero_destinacion_sql import (
    parse_destinacion_subcadenas,
    sql_destinacion_import_or,
)
from tasa_estadistica.export.recupero_formulas import build_recupero_v2_row_with_formulas

# Cabeceras fila 3 del modelo «V2_Ejemplo» (25 columnas).
RECUPERO_V2_HEADERS: tuple[str, ...] = (
    "PROVEEDOR",
    "D.I.",
    "OFICIALIZACION",
    "TC DESPACHO",
    "ALICUOTA COBRADA",
    "T.E. 011",
    "T.E. MONT MAX 061",
    "T.E. MONT MAX2 062",
    "T.E. TOTAL COBRADA",
    "BASE RECONSTR. S/011",
    "FOB",
    "FLETE",
    "SEGURO",
    "CIF DOCUMENTAL",
    "DIF. BASE VS CIF",
    "T.E. CORRECTA S/BASE",
    "T.E. CORRECTA S/CIF",
    "EXCESO USD S/BASE",
    "EXCESO USD S/CIF",
    "A SOLICITAR ARS S/BASE",
    "A SOLICITAR ARS S/CIF",
    "TIENE MONT MAX",
    "METODO SUGERIDO",
    "ESTADO REVISION",
    "ARCHIVO FUENTE",
)


def recupero_v2_excel_currency_format(c: int) -> str | None:
    """
    Formato de celda numérica en la hoja recupero_V2 (columna Excel 1-based, fila datos).

    Alineado con `tasa_estadistica.web.recupero_display` (USD vs ARS por nombre de columna).
    """
    if c == 3:
        return "fecha"
    if c == 4:
        return "tc"
    if c == 5:
        return "alic"
    if c in (6, 7, 8, 9, 20, 21):
        return "ars"
    if 10 <= c <= 19:
        return "usd"
    return None


# Colores cabecera fila (1-based): qué contrastar con AFIP vs fórmulas en tu modelo.
_FILL_AFIP_PRIMERA_FILA = PatternFill(
    "solid", fgColor="C6EFCE"
)  # verde: PROVEEDOR … T.E. 062 (imagen 1)
_FILL_AFIP_TOTAL_LIQ = PatternFill(
    "solid", fgColor="E2EFDA"
)  # verde claro: total 011+061+062 desde liquidación
_FILL_AFIP_FOB_FLETE_SEG = PatternFill(
    "solid", fgColor="C6EFCE"
)  # verde: FOB, FLETE, SEGURO (imagen 2)
_FILL_FORMULAS = PatternFill("solid", fgColor="F2F2F2")  # gris: resto = fórmulas en tu Excel


def _style_recupero_headers(ws: Worksheet, header_row: int = 3) -> None:
    """Marca visualmente columnas «dato AFIP» vs «fórmula» (según tu modelo)."""
    n = len(RECUPERO_V2_HEADERS)
    bold = Font(bold=True)
    for c in range(1, n + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = bold
        if 1 <= c <= 8:
            cell.fill = _FILL_AFIP_PRIMERA_FILA
        elif c == 9:
            cell.fill = _FILL_AFIP_TOTAL_LIQ
        elif c == 10:
            cell.fill = _FILL_FORMULAS
        elif 11 <= c <= 13:
            cell.fill = _FILL_AFIP_FOB_FLETE_SEG
        else:
            cell.fill = _FILL_FORMULAS


def _fill_como_verificar_sheet(ws: Worksheet) -> None:
    ws.title = "como_verificar"
    lines = [
        ("Cómo verificar que los datos coinciden con AFIP", ""),
        ("", ""),
        (
            "1) Misma referencia",
            "Usá el mismo D.I. (columna B) en ARCA/AFIP que en este archivo.",
        ),
        (
            "2) Liquidación (lo que ya extraemos del WS de liquidaciones)",
            "OFICIALIZACION = fecha de la liquidación; TC DESPACHO = cotización del resumen; "
            "T.E. 011 / 061 / 062 = importes del concepto con ese código (MontoPagado). "
            "Compará con el detalle de la liquidación en el sitio AFIP/ARCA para ese despacho.",
        ),
        (
            "3) Total T.E.",
            "T.E. TOTAL COBRADA aquí es la suma 011+061+062 de esa liquidación (control rápido).",
        ),
        (
            "4) PROVEEDOR",
            "Si el fetch fue por MOA (wconsdeclaracion) y se guardó declaracion_listado, "
            "se rellena desde esa fila (campos típicos DenominacionProveedorExterior, etc.). "
            "Si la base se cargó sin ese bloque, hacé un fetch de nuevo o completá manual.",
        ),
        (
            "5) ALICUOTA, FOB, FLETE, SEGURO",
            "Si AFIP/MOA los envía en el JSON de la liquidación, se copian. "
            "CIF DOCUMENTAL se calcula como FOB+FLETE+SEGURO (igual que en el Excel modelo).",
        ),
        (
            "6) Columnas grises en la hoja recupero_V2",
            "Excedentes, «A solicitar», métodos sugeridos, etc.; "
            "suelen ser fórmulas en tu archivo.",
        ),
        (
            "7) Tu Excel manual",
            "Podés abrir en paralelo «Modelo Excel - Recupero…» y comparar fila a fila "
            "las celdas verdes para el mismo D.I.",
        ),
    ]
    for row in lines:
        ws.append(row)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 85


def _latest_ic_liquidacion_ids(
    conn: sqlite3.Connection, desde: date, hasta: date, settings: Settings
) -> list[int]:
    """
    Un `liquidaciones.id` por `destinacion_id`.

    Puede haber **dos filas** para el mismo despacho (MOA wconsdeclaracion + complemento
    `consultarLiquidaciones`): `id_externo` distinto → no deduplican al guardar, y el
    `MAX(id)` quedaba en la fila Zeep **sin** `moa_detallada_*` → FOB/FLETE/SEGURO vacíos.
    Preferimos la fila cuyo `raw_json` incluye datos MOA; si ninguna, el `id` más alto.
    """
    cond, extras = sql_destinacion_import_or(
        "l2", parse_destinacion_subcadenas(settings)
    )
    # Coincidencia literal en JSON serializado por `json.dumps` (claves con comillas dobles).
    moa_prio = (
        "CASE WHEN l2.raw_json LIKE '%\"moa_detallada_caratula\"%' "
        "OR l2.raw_json LIKE '%\"moa_detallada_liquidaciones_detalle\"%' "
        "THEN 1 ELSE 0 END"
    )
    sql = f"""
        SELECT sub.id FROM (
            SELECT l2.id,
                   ROW_NUMBER() OVER (
                       PARTITION BY l2.destinacion_id
                       ORDER BY {moa_prio} DESC, l2.id DESC
                   ) AS rn
            FROM liquidaciones l2
            WHERE l2.fecha >= ?
              AND l2.fecha <= ?
              AND {cond}
        ) AS sub
        WHERE sub.rn = 1
        ORDER BY sub.id
        """
    cur = conn.execute(sql, (desde.isoformat(), hasta.isoformat(), *extras))
    return [int(r[0]) for r in cur.fetchall() if r[0] is not None]


def _stats_liquidaciones_rango(
    conn: sqlite3.Connection, desde: date, hasta: date
) -> tuple[int, int]:
    """Total filas `liquidaciones` en rango y destinaciones distintas no vacías."""
    fd, fh = desde.isoformat(), hasta.isoformat()
    n_liq = int(
        conn.execute(
            "SELECT COUNT(*) FROM liquidaciones WHERE fecha >= ? AND fecha <= ?",
            (fd, fh),
        ).fetchone()[0]
    )
    n_dest = int(
        conn.execute(
            """
            SELECT COUNT(DISTINCT destinacion_id) FROM liquidaciones
            WHERE fecha >= ? AND fecha <= ?
              AND destinacion_id IS NOT NULL AND TRIM(destinacion_id) != ''
            """,
            (fd, fh),
        ).fetchone()[0]
    )
    return n_liq, n_dest


def _sum_te_011_061_062(
    conn: sqlite3.Connection, liquidacion_id: int
) -> tuple[Decimal, Decimal, Decimal]:
    cur = conn.execute(
        """
        SELECT codigo, importe, raw_json
        FROM conceptos_liquidacion
        WHERE liquidacion_id = ?
        """,
        (liquidacion_id,),
    )
    s011 = s061 = s062 = Decimal("0")
    for row in cur.fetchall():
        code = (row["codigo"] or "").strip().upper()
        amt = importe_concepto_efectivo(row["importe"], row["raw_json"])
        if code == "011":
            s011 += amt
        elif code == "061":
            s061 += amt
        elif code == "062":
            s062 += amt
    return s011, s061, s062


def _excel_num(d: Decimal) -> float | int:
    if d == d.to_integral():
        return int(d)
    return float(d)


def _apply_recupero_number_formats(ws: Worksheet, data_start_row: int, n_rows: int) -> None:
    """Formatos de moneda en Excel alineados al panel: $ (ARS) y U$S (USD)."""
    if n_rows <= 0:
        return
    fmt_ars = '"$" #,##0.00'
    fmt_usd = '"U$S" #,##0.00'
    fmt_tc = "#,##0.00"
    fmt_alic = "0.00"
    fmt_fecha = "DD/MM/YYYY"
    end_row = data_start_row + n_rows - 1
    ncols = len(RECUPERO_V2_HEADERS)
    fmt_by_kind = {
        "fecha": fmt_fecha,
        "tc": fmt_tc,
        "alic": fmt_alic,
        "ars": fmt_ars,
        "usd": fmt_usd,
    }
    for r in range(data_start_row, end_row + 1):
        for c in range(1, ncols + 1):
            kind = recupero_v2_excel_currency_format(c)
            if kind is None:
                continue
            cell = ws.cell(row=r, column=c)
            cell.number_format = fmt_by_kind[kind]


def _override_row(conn: sqlite3.Connection, destinacion_id: str) -> dict[str, str]:
    """Lee override manual si la tabla existe (no rompe si faltase)."""
    if not (destinacion_id or "").strip():
        return {}
    try:
        row = conn.execute(
            "SELECT fob, flete, seguro, nota FROM recupero_overrides WHERE destinacion_id = ?",
            (destinacion_id.strip(),),
        ).fetchone()
    except sqlite3.OperationalError:
        return {}
    if row is None:
        return {}
    return {
        "fob": (row["fob"] or "").strip() if hasattr(row, "keys") else (row[0] or "").strip(),
        "flete": (row["flete"] or "").strip() if hasattr(row, "keys") else (row[1] or "").strip(),
        "seguro": (row["seguro"] or "").strip() if hasattr(row, "keys") else (row[2] or "").strip(),
        "nota": (row["nota"] or "").strip() if hasattr(row, "keys") else (row[3] or "").strip(),
    }


def _extra_recupero_unificado_por_destinacion(
    conn: sqlite3.Connection,
    destinacion_id: str,
    primary_raw: str | bytes | None,
) -> dict[str, str]:
    """
    Precedencia de FOB/FLETE/SEGURO (y alícuota/base):

      1. Override manual (`recupero_overrides`) si tiene el valor cargado.
      2. Fila MOA primaria (la de mayor id con carátula para ese D.I.).
      3. Cualquier otra `liquidaciones` con el mismo D.I. (rellena huecos si (2) vino vacío).

    El paso 3 permite que un refetch que dejó la fila nueva con MOA vacío siga funcionando
    gracias a una fila previa con carátula completa. Los overrides siempre ganan para que el
    usuario pueda corregir datos que AFIP no informa (declaraciones CANC, etc.).
    """
    merged = extract_recupero_valores_extra(primary_raw)
    fill_keys = ("fob", "flete", "seguro", "alicuota", "base_reconstr")

    override = _override_row(conn, destinacion_id)
    for k in ("fob", "flete", "seguro"):
        if (override.get(k) or "").strip():
            merged[k] = override[k]

    if not (destinacion_id or "").strip():
        return merged
    if all((merged.get(k) or "").strip() for k in ("fob", "flete", "seguro")):
        return merged
    cur = conn.execute(
        """
        SELECT raw_json FROM liquidaciones
        WHERE destinacion_id = ? AND TRIM(destinacion_id) != ''
        ORDER BY id DESC
        """,
        (destinacion_id.strip(),),
    )
    for (rj,) in cur.fetchall():
        cand = extract_recupero_valores_extra(rj)
        for k in fill_keys:
            if not (merged.get(k) or "").strip() and (cand.get(k) or "").strip():
                merged[k] = cand[k]
        if all((merged.get(k) or "").strip() for k in ("fob", "flete", "seguro")):
            break
    return merged


def recupero_v2_data_rows(
    conn: sqlite3.Connection,
    fecha_desde: date,
    fecha_hasta: date,
    settings: Settings | None = None,
    *,
    out_liquidacion_ids: list[int] | None = None,
) -> list[list[Any]]:
    """Filas de datos (sin cabecera) para recupero V2: una por despacho (subcadenas en settings)."""
    settings = settings or get_settings()
    rows: list[list[Any]] = []
    for lid in _latest_ic_liquidacion_ids(conn, fecha_desde, fecha_hasta, settings):
        row = conn.execute(
            "SELECT destinacion_id, fecha, raw_json FROM liquidaciones WHERE id = ?",
            (lid,),
        ).fetchone()
        if not row:
            continue
        dest = row["destinacion_id"] or ""
        fecha_s = row["fecha"]
        raw_json = row["raw_json"]
        fl = flatten_liquidacion_raw_json(raw_json)
        cot = fl.get("liq_Cotizacion") or ""
        try:
            tc = float(str(cot).replace(",", ".").strip()) if str(cot).strip() else None
        except ValueError:
            tc = None

        te011, te061, te062 = _sum_te_011_061_062(conn, lid)

        oficial: date | None = None
        if fecha_s:
            try:
                oficial = date.fromisoformat(str(fecha_s)[:10])
            except ValueError:
                oficial = None

        proveedor = extract_proveedor_from_liquidacion_raw_json(raw_json)
        extra = _extra_recupero_unificado_por_destinacion(conn, dest, raw_json)
        fob_s = extra.get("fob") or ""
        flete_s = extra.get("flete") or ""
        seguro_s = extra.get("seguro") or ""

        rows.append(
            build_recupero_v2_row_with_formulas(
                proveedor=proveedor,
                destinacion_id=dest,
                oficial=oficial,
                tc_raw=tc if tc is not None else "",
                alicuota_cobrada_str=extra.get("alicuota") or "",
                base_reconstr_str=extra.get("base_reconstr") or "",
                te011=te011,
                te061=te061,
                te062=te062,
                fob_str=fob_s,
                flete_str=flete_s,
                seguro_str=seguro_s,
            )
        )
        if out_liquidacion_ids is not None:
            out_liquidacion_ids.append(lid)
    return rows


def _fila_fob_flete_seguro_alguno_vacio(row: list[Any]) -> bool:
    for col in ("FOB", "FLETE", "SEGURO"):
        i = RECUPERO_V2_HEADERS.index(col)
        v = row[i] if i < len(row) else ""
        if v == "" or v is None:
            return True
    return False


def _avisos_fob_recupero_panel(
    conn: sqlite3.Connection,
    lids: list[int],
    data_rows: list[list[Any]],
) -> list[str]:
    """
    Textos cortos para el panel cuando hay despachos sin montos USD en la grilla.
    No sustituye datos: solo explica límites de AFIP/MOA (p. ej. declaración CANC).
    """
    avisos: list[str] = []
    for lid, row in zip(lids, data_rows, strict=True):
        if not _fila_fob_flete_seguro_alguno_vacio(row):
            continue
        r = conn.execute(
            "SELECT destinacion_id, raw_json FROM liquidaciones WHERE id = ?",
            (lid,),
        ).fetchone()
        if not r:
            continue
        dest = (r["destinacion_id"] or "").strip() or "(sin D.I.)"
        raw_s = r["raw_json"] or ""
        decl: dict[str, Any] = {}
        try:
            top = json.loads(raw_s) if raw_s else {}
            inner = top.get("raw") if isinstance(top, dict) else {}
            if isinstance(inner, dict):
                dlist = inner.get("declaracion_listado")
                if isinstance(dlist, dict):
                    decl = dlist
        except json.JSONDecodeError:
            pass
        estado = str(
            decl.get("CodigoEstadoDeclaracion")
            or decl.get("codigoEstadoDeclaracion")
            or ""
        ).strip().upper()
        if estado == "CANC":
            avisos.append(
                f"Despacho {dest}: en MOA la declaración está CANC (cancelada). "
                "AFIP suele dejar de informar FOB/FLETE/SEGURO en listado/detalle para ese estado; "
                "si necesitás esos montos, revisá en ARCA o usá un despacho no cancelado."
            )
        elif '"moa_detallada_caratula"' not in raw_s and _fila_fob_flete_seguro_alguno_vacio(row):
            avisos.append(
                f"Despacho {dest}: en SQLite no hay carátula MOA (`moa_detallada_caratula`). "
                "Volvé a ejecutar `tasa-arca fetch` para ese período con el código actual."
            )
    return avisos


def recupero_resumen_local(
    db_path: Path,
    fecha_desde: date,
    fecha_hasta: date,
    settings: Settings | None = None,
) -> dict[str, Any]:
    """
    Totales desde SQLite: despachos que pasan el filtro de importación en rango
    y suma T.E. 011+061+062 (ARS liquidación).
    Útil para panel web / demo sin depender de red.
    """
    settings = settings or get_settings()
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        n_liq, n_dest_all = _stats_liquidaciones_rango(conn, fecha_desde, fecha_hasta)
        lids_usadas: list[int] = []
        rows = recupero_v2_data_rows(
            conn, fecha_desde, fecha_hasta, settings, out_liquidacion_ids=lids_usadas
        )
        avisos_fob = _avisos_fob_recupero_panel(conn, lids_usadas, rows)
        # D.I. con override cargado (para que el panel los marque y permita editarlos).
        try:
            dest_ids_presentes = [
                (row[RECUPERO_V2_HEADERS.index("D.I.")] or "").strip()
                for row in rows
            ]
            dest_ids_presentes = [d for d in dest_ids_presentes if d]
            if dest_ids_presentes:
                placeholders = ",".join("?" * len(dest_ids_presentes))
                cur = conn.execute(
                    f"SELECT destinacion_id, fob, flete, seguro, nota "
                    f"FROM recupero_overrides WHERE destinacion_id IN ({placeholders})",
                    dest_ids_presentes,
                )
                overrides_activos = {
                    r["destinacion_id"]: {
                        "fob": r["fob"] or "",
                        "flete": r["flete"] or "",
                        "seguro": r["seguro"] or "",
                        "nota": r["nota"] or "",
                    }
                    for r in cur.fetchall()
                }
            else:
                overrides_activos = {}
        except sqlite3.OperationalError:
            overrides_activos = {}
    finally:
        conn.close()
    total = Decimal("0")
    for r in rows:
        if len(r) <= 8:
            continue
        v = r[8]
        if v == "" or v is None:
            continue
        if isinstance(v, Decimal):
            total += v
        elif isinstance(v, (int, float)):
            total += Decimal(str(v))
        else:
            try:
                total += Decimal(str(v).replace(",", "."))
            except Exception:
                pass
    return {
        "n_despachos": len(rows),
        "total_tasa_recupero_ars": total,
        "filas": rows,
        "n_liquidaciones_en_rango": n_liq,
        "n_destinaciones_distintas_en_rango": n_dest_all,
        "destinacion_subcadenas": ",".join(parse_destinacion_subcadenas(settings)),
        "recupero_avisos_fob": avisos_fob,
        "recupero_overrides": overrides_activos,
    }


def write_recupero_v2_csv(output_path: Path, data_rows: list[list[Any]]) -> Path:
    """CSV UTF-8 con BOM: misma presentación que el panel ($ / U$S, fechas AR)."""
    from tasa_estadistica.web.recupero_display import format_recupero_cell_safe

    output_path.parent.mkdir(parents=True, exist_ok=True)
    keys = list(RECUPERO_V2_HEADERS)
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(keys)
        for row in data_rows:
            cells = [
                format_recupero_cell_safe(col, row[i] if i < len(row) else "")
                for i, col in enumerate(keys)
            ]
            w.writerow(cells)
    return output_path


def build_recupero_v2_excel(
    db_path: Path,
    output_path: Path,
    fecha_desde: date,
    fecha_hasta: date,
    settings: Settings | None = None,
) -> Path:
    """
    Genera un .xlsx con una hoja `recupero_V2` alineada al modelo manual.
    Rellena desde SQLite: D.I., fecha, TC (cotización), T.E. 011/061/062 y total.
    Si en `raw_json` hay `declaracion_listado` / resumen AFIP con FOB, CIF, alícuota, etc.,
    también se copian a las columnas correspondientes; si no vinieron en el fetch, quedan vacías.
    """
    settings = settings or get_settings()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row

    data_rows = recupero_v2_data_rows(conn, fecha_desde, fecha_hasta, settings)

    wb = Workbook()
    _fill_como_verificar_sheet(wb.active)
    ws = wb.create_sheet("recupero_V2")

    ws.append(["Fuente datos", "SQLite liquidaciones + conceptos (AFIP)"])
    ws.append(
        [
            "Leyenda cabecera (fila 3)",
            "Verde = dato a contrastar con AFIP (imágenes). Gris = fórmulas en tu modelo. "
            "Verde claro (T.E. TOTAL) = suma AFIP 011+061+062.",
        ]
    )
    ws.append(list(RECUPERO_V2_HEADERS))
    _style_recupero_headers(ws, header_row=3)

    for out_row in data_rows:
        ws.append(out_row)

    _apply_recupero_number_formats(ws, data_start_row=4, n_rows=len(data_rows))

    conn.close()

    for i in range(1, len(RECUPERO_V2_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    wb.save(output_path)
    write_recupero_v2_csv(output_path.with_suffix(".csv"), data_rows)
    return output_path
